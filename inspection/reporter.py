"""
inspection.reporter — Excel Report Generator
=============================================
Generates a formatted Excel workbook with:
  - Summary sheet: one row per device with all inspection results
  - Per-device Interface sheets: detailed interface status

Uses openpyxl for conditional formatting and styling.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    NamedStyle, numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table
from openpyxl.formatting.rule import CellIsRule


class ExcelReporter:
    """Generate inspection report in Excel format."""

    # Color definitions
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    NORMAL_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def __init__(self):
        self.logger = logging.getLogger("inspection.reporter")

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------
    def generate(self, all_results: dict, output_path: str) -> None:
        """
        Generate the full Excel report.

        Args:
            all_results: {host_name: {module_name: result_dict}}
            output_path: Path for the output .xlsx file.
        """
        wb = openpyxl.Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create Summary sheet
        ws_summary = wb.create_sheet("汇总 Summary", 0)
        self._write_summary(ws_summary, all_results)

        # Create per-device Interface sheets
        for host_name, modules in all_results.items():
            iface_data = modules.get("interfaces", {}).get("interfaces", [])
            if iface_data:
                ws_iface = wb.create_sheet(f"{host_name[:28]} 接口")
                self._write_interface_sheet(ws_iface, host_name, iface_data)

        # Save
        wb.save(output_path)
        self.logger.info("Excel report saved to %s", output_path)

    # ------------------------------------------------------------------
    # Summary Sheet
    # ------------------------------------------------------------------
    def _write_summary(self, ws, all_results: dict) -> None:
        """Write the Summary sheet with all inspection results."""

        headers = [
            "设备名称", "管理 IP", "厂商", "角色", "位置",
            "可达性", "CPU 使用率", "内存使用率",
            "风扇状态", "电源状态", "温度状态",
            "接口 UP 数", "接口 DOWN 数", "告警接口数",
            "配置备份", "巡检时间",
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self.THIN_BORDER

        # Write data rows
        row = 2
        for host_name, modules in all_results.items():
            # Gather data from all modules
            ping = modules.get("ping", {})
            cpu_mem = modules.get("cpu_memory", {})
            hw = modules.get("hardware", {})
            iface = modules.get("interfaces", {})
            backup = modules.get("config_backup", {})

            # Extract host info
            host_info = self._get_host_info(host_name, modules)

            # NTP column value
            ntp_alert = ntp.get("ntp_alert", "normal")
            ntp_display = "同步" if ntp.get("ntp_synchronized") else "未同步"
            # Compliance column value
            comp_alert = compliance.get("compliance_alert", "normal")
            comp_pass = compliance.get("pass_count", 0)
            comp_fail = compliance.get("fail_count", 0)
            comp_display = f"通过{comp_pass}/失败{comp_fail}" if comp_pass or comp_fail else "-"
            # Log check column value
            log_alert = log_check.get("log_alert", "normal")
            log_matches = log_check.get("total_matches", 0)
            log_display = f"{log_matches}条异常" if log_matches else "正常"

            data_row = [
                host_name,
                host_info.get("ip", ""),
                host_info.get("vendor", ""),
                host_info.get("role", ""),
                host_info.get("site", ""),
                "可达" if ping.get("reachable") else "不可达",
                self._fmt_pct(cpu_mem.get("cpu_usage_pct")),
                self._fmt_pct(cpu_mem.get("memory_usage_pct")),
                self._fmt_alert(hw.get("fan_alert", "normal")),
                self._fmt_alert(hw.get("power_alert", "normal")),
                self._fmt_alert(hw.get("temp_alert", "normal")),
                iface.get("total_up", 0),
                iface.get("total_down", 0),
                len(iface.get("alert_interfaces", [])),
                "已备份" if backup.get("backup_status") == "ok" else "失败",
                ntp_display,
                comp_display,
                log_display,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ]

            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.THIN_BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Apply conditional coloring
            self._color_summary_row(ws, row, cpu_mem, hw, iface, backup)

            row += 1

        # Auto-fit column widths
        self._auto_fit_columns(ws, len(headers))

        # Freeze header row
        ws.freeze_panes = "A2"

    def _get_host_info(self, host_name: str, modules: dict) -> dict:
        """Extract host metadata from module results."""
        info = {
            "ip": host_name,
            "vendor": "",
            "role": "",
            "site": "",
        }
        # Try to get vendor from ping result
        ping = modules.get("ping", {})
        if ping.get("vendor"):
            info["vendor"] = ping["vendor"]

        return info

    def _color_summary_row(self, ws, row: int,
                           cpu_mem: dict, hw: dict,
                           iface: dict, backup: dict) -> None:
        """Apply per-cell coloring based on alert levels."""

        # CPU (column 7)
        cpu_alert = cpu_mem.get("cpu_alert", "normal")
        ws.cell(row=row, column=7).fill = self._alert_fill(cpu_alert)

        # Memory (column 8)
        mem_alert = cpu_mem.get("memory_alert", "normal")
        ws.cell(row=row, column=8).fill = self._alert_fill(mem_alert)

        # Fan (column 9)
        ws.cell(row=row, column=9).fill = self._alert_fill(hw.get("fan_alert", "normal"))

        # Power (column 10)
        ws.cell(row=row, column=10).fill = self._alert_fill(hw.get("power_alert", "normal"))

        # Temperature (column 11)
        ws.cell(row=row, column=11).fill = self._alert_fill(hw.get("temp_alert", "normal"))

        # Interface down count (column 13) — red if > 0
        if iface.get("total_down", 0) > 0:
            ws.cell(row=row, column=13).fill = self.RED_FILL

        # Alert interfaces (column 14) — red if > 0
        if len(iface.get("alert_interfaces", [])) > 0:
            ws.cell(row=row, column=14).fill = self.RED_FILL

        # Backup (column 15) — red if failed
        if backup.get("backup_status") != "ok":
            ws.cell(row=row, column=15).fill = self.RED_FILL

    # ------------------------------------------------------------------
    # Interface Sheet
    # ------------------------------------------------------------------
    def _write_interface_sheet(self, ws, host_name: str,
                               interfaces: list[dict]) -> None:
        """Write per-device interface detail sheet."""

        headers = [
            "接口名称", "状态", "速率 (Mbps)", "MTU",
            "入方向错包", "出方向错包", "CRC 错包", "告警",
        ]

        # Header row
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.THIN_BORDER

        # Data rows
        for i, iface in enumerate(interfaces, 2):
            data = [
                iface.get("name", ""),
                iface.get("status", "down"),
                iface.get("speed", 0),
                iface.get("mtu", 0),
                iface.get("in_errors", 0),
                iface.get("out_errors", 0),
                iface.get("in_crc_errors", 0),
                "是" if iface.get("alert") else "否",
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=i, column=col, value=value)
                cell.border = self.THIN_BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Color coding
            if iface.get("status") != "up":
                ws.cell(row=i, column=2).fill = self.RED_FILL

            if iface.get("alert"):
                ws.cell(row=i, column=8).fill = self.RED_FILL

            if iface.get("in_crc_errors", 0) > 0:
                ws.cell(row=i, column=7).fill = self.YELLOW_FILL

        # Auto-fit
        self._auto_fit_columns(ws, len(headers))
        ws.freeze_panes = "A2"

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    @staticmethod
    def _fmt_pct(value) -> str:
        """Format a numeric value as percentage string."""
        if value is None:
            return "N/A"
        try:
            return f"{float(value):.1f}%"
        except (ValueError, TypeError):
            return str(value)

    @staticmethod
    def _fmt_alert(alert: str) -> str:
        """Translate alert level to Chinese label."""
        mapping = {
            "normal": "正常",
            "warning": "预警",
            "critical": "告警",
        }
        return mapping.get(alert, alert)

    @staticmethod
    def _alert_fill(alert: str) -> PatternFill:
        """Return the appropriate fill for an alert level."""
        mapping = {
            "normal": ExcelReporter.GREEN_FILL,
            "warning": ExcelReporter.YELLOW_FILL,
            "critical": ExcelReporter.RED_FILL,
        }
        return mapping.get(alert, ExcelReporter.NORMAL_FILL)

    @staticmethod
    def _auto_fit_columns(ws, num_cols: int) -> None:
        """Auto-fit column widths based on content."""
        for col in range(1, num_cols + 1):
            max_length = 0
            col_letter = get_column_letter(col)
            for row in ws.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    if cell.value:
                        # Estimate width: CJK characters count as 2
                        val = str(cell.value)
                        length = sum(2 if ord(c) > 127 else 1 for c in val)
                        max_length = max(max_length, length)
            # Cap width
            adjusted = min(max_length + 4, 40)
            ws.column_dimensions[col_letter].width = max(adjusted, 10)
